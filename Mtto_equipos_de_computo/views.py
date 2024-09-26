from django.shortcuts import render, redirect
from .Signatures import Proces_signatures
from .forms import ComputerEquipmentMaintenanceForm
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.drawing.image import Image as OpenpyxlImage
import shutil
from PIL import Image as PILImage

date_str = datetime.now().strftime('%Y-%m-%d')

def Maintenance (request):
    if request.method == 'POST':
        form = ComputerEquipmentMaintenanceForm(request.POST)
        if form.is_valid():
            Equipment_data = form.cleaned_data
            new_file_name = f"Z:/SERVER/Dagelec/Mantenimiento/03. EQUIPOS INFORMATICOS/10. MANTENIMIENTO/{str(Equipment_data['computer_equipment_id'])}_{date_str}.xlsx"
            shutil.copy('C:/Users/Dagelec LTDA/Desktop/Eco_forms/Eco_forms/Pruebas_excel/Plantilla_Mtto_de_computo.xlsx', new_file_name)

            wb = load_workbook(filename= new_file_name)
            ws = wb.active

            # Asignar valores a las celdas correspondientes
            ws['B7'] = Equipment_data['computer_equipment_id']
            ws['B8'] = Equipment_data['inspection_date']
            ws['B9'] = Equipment_data['brand']
            ws['T7'] = 'X' if Equipment_data['maintenance_type'] == 'Preventivo' else ''
            ws['X7'] = 'X' if Equipment_data['maintenance_type'] == 'Correctivo' else ''
            ws['T9'] = 'X' if Equipment_data['equipment_type'] == 'Escritorio' else ''
            ws['X9'] = 'X' if Equipment_data['equipment_type'] == 'Laptop' else ''

            # Campos de inspección visual
            inspections = [
                'visual_inspection', 'complete_screws', 'charging_port', 'port_verification',
                'keyboard_check', 'touchpad_check', 'screen_check', 'camera_microphone_verification',
                'monitors', 'peripherals', 'battery_check', 'charger_check'
            ]

            row_start = 12
            for i, inspection in enumerate(inspections):
                row = row_start + i
                ws[f'J{row}'] = 'X' if Equipment_data[inspection] == 'Cumple' else ''
                ws[f'K{row}'] = 'X' if Equipment_data[inspection] == 'No cumple' else ''
                ws[f'L{row}'] = 'X' if Equipment_data[inspection] == 'No aplica' else ''
                ws[f'M{row}'] = request.POST.get(f'observaciones_{inspection}', '')
                ws[f'S{row}'] = request.POST.get(f'acciones_{inspection}', '')

            os_antivirus_inspections = [
                'os_verification', 'antivirus_verification', 'hard_drive_status', 'temp_files_cleanup',
                'pending_updates', 'license_verification'
            ]

            row_start = 26
            for i, inspection in enumerate(os_antivirus_inspections):
                row = row_start + i
                ws[f'J{row}'] = 'X' if Equipment_data[inspection] == 'Cumple' else ''
                ws[f'K{row}'] = 'X' if Equipment_data[inspection] == 'No cumple' else ''
                ws[f'L{row}'] = 'X' if Equipment_data[inspection] == 'No aplica' else ''
                ws[f'M{row}'] = request.POST.get(f'observaciones_{inspection}', '')
                ws[f'S{row}'] = request.POST.get(f'acciones_{inspection}', '')

            # Otros campos
            ws['A35'] = Equipment_data['corrective_maintenance_notes']
            ws['A43'] = Equipment_data['general_observations']
            Cell_value = ws['A52'].value
            ws['A52'] = f"{Cell_value} {Equipment_data['received_by']}"
            Cell_value = ws['K52'].value
            ws['K52'] = f"{Cell_value} {Equipment_data['executed_by']}"

            if ';base64,' in request.POST['firmaData'] and ';base64,' in request.POST['firmaData2']:
                firma1_path, firma2_path = Proces_signatures(request)
            else:
                Error = 'Debes firmar todos los campos'
                return render(request, 'Forms/FormG.html', {'form': form,
                                                            'Error': Error
                                                            })
            def resize_image(image_path, width, height):
                with PILImage.open(image_path) as img:
                    img = img.resize((width, height), PILImage.LANCZOS)
                    img.save(image_path)

            # Redimensionar las imágenes a 130x50 píxeles
            resize_image(firma1_path, 200, 75)
            resize_image(firma2_path, 200, 75)

             # Insertar las imágenes en el archivo de Excel
            img1 = OpenpyxlImage(firma1_path)
            img2 = OpenpyxlImage(firma2_path)
            ws.add_image(img1, 'A56')
            ws.add_image(img2, 'K56')

            # Guardar el archivo de Excel
            wb.save(new_file_name)
            return redirect('Index')
    else:
        form = ComputerEquipmentMaintenanceForm()
        Titulo = "Mantenimiento de equipos de computo"
        return render(request, 'Form.html', {'form': form,
                                             'Title' : Titulo})