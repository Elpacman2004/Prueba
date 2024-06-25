from openpyxl import load_workbook
from django.http import HttpRequest

def Cell_selectorE (Name_file, form):
    Cell_assignment = {
        "cmc_356_omicron": "H15",
        "cmc_353_omicron": "H16",
        "doble_f6150": "H17",
        "ponovo_l336i": "H18",
        "cpc_100_omicron": "H19",
        "cp_td1_omicron": "H20",
        "cp_cu1_omicron": "H21",
        "cp_cr500_omicron": "H22",
        "ct_analyzer_omicron": "H23",
        "franalizer_omicron": "H24",
        "cp_cb2_omicron": "H25",
        "ponovo_pct_200ai": "H26",
        "factor_hv_hipot_gd6000a": "H27",
        "potencia_tang_delta_avo": "H28",
        "micro_contactos_dmo200": "H29",
        "vacuometro_digital": "H30",
        "calidad_energia_metrel": "R15",
        "tiempos_operacion_gdkc_6a": "R16",
        "tiempos_operacion_egil": "R17",
        "sata40_tension_minima": "R18",
        "relacion_trans_ttr_gbii": "R19",
        "r_devanados_hv_gdzrs_20a": "R20",
        "r_devanados_hv_gdzrc_5a": "R21",
        "milliohm_sonel": "R22",
        "medidor_mit520_megger": "R23",
        "medidor_mit500_megger": "R24",
        "medidor_mic_5010_sonel": "R25",
        "medidor_mic_5015_sonel": "R26",
        "medidor_mic_5010_sonel_25kv": "R27",
        "medidor_mic_5005_sonel": "R28",
        "telurometro_metrotech": "R29",
        "chisporroteo_megger_ots60bp": "R30",
        "circuito_mpc_5_5_50": "AB16",
        "probador_cts_cct_xsaid": "AB15",
        "vaisala_punto_rocio": "AB17",
        "recuperador_gas_sf6_gr": "AB18",
        "recuperador_gas_sf6_pe": "AB19",
        "bomba_vacio_sf6": "AB20",
        "analizador_gas_sf6_grande": "AB21",
        "analizador_gas_sf6_pequeno": "AB22",
        "kit_llenado_sf6": "AB23",
        "kit_racores": "AB24",
        "pd_scan_descargas_parciales": "AB25",
        "camara_termografica_nec": "AB26",
        "camara_termografica_irys": "AB27",
        "fuentes_110_220v": "AB29",
        "secuencimetro": "AB28"
    }
    wb = load_workbook(filename=Name_file)
    sheet = wb['Hoja1']
    
    for field, value in form.items():
        if field in Cell_assignment:
            cell = Cell_assignment[field]
            sheet[cell] = value
            
    wb.save(Name_file)
    
def Cell_selectorT(Name_file, form):
    Cell_assignment = {
  "planta_electrica" : 'H35',
  "juego_de_tierras": 'H36',
  "caja_herramientas": 'H37',
  "pertiga_escopeta": 'H38',
  "pertiga_telescopica": 'H39',
  "pistola_calor": 'H40',
  "motor_tool": 'H41',
  "kit_soldadura_cautin": 'H42',
  "pistola_silicona": 'H43',
  "sopladora": 'H44',
  "taladro": 'H45',
  "filtro_sf6": 'H46',
  "compresor_pintura": 'H47',
  "compresor_aire_seco": 'R35',
  "equipo_soldadura": 'R36',
  "radio_transmissor": 'R37',
  "higrometro": 'R38',
  "pata_de_cabra": 'R39',
  "ponchadora_hidraulica": 'R40',
  "ponchadora_neumatica": 'R41',
  "cizalla_manual": 'R42',
  "tarrajero": 'R43',
  "bomba_para_aceite": 'R44',
  "pistola_pintura": 'R45',
  "extension_electrica": 'R46',
  "esmeril": 'R47',
  "sonda": 'AB35',
  "kit_muestra_aceite": 'AB36',
  "tarraja": 'AB37',
  "aspiradora": 'AB38',
  "pulidora": 'AB39',
  "caladora": 'AB40',
  "reflector": 'AB41',
  "tijera_metal": 'AB42',
  "poleas": 'AB43',
  "maleta_sf6": 'AB44',
  "maquina_filtroprensa": 'AB45'
    } 
    wb = load_workbook(filename=Name_file)
    sheet = wb['Hoja1']
    
    for field, value in form.items():
        if field in Cell_assignment:
            cell = Cell_assignment[field]
            sheet[cell] = value
            
    wb.save(Name_file)
    
def Cell_selectorS(Name_file, form):
    Cell_assignment = {
        "botiquin" : 'H52', 
        "camilla" : 'H53',
        "colombinas" : 'H54',
        "conos" : 'H55',
        "candados_seguridad" : 'H56',
        "detector_tension" : 'H57',
        "detector_fugas_sf6" : 'H58',
        "arnes" : 'R52',
        "linea_vida" : 'R53',
        "extintor_multiproposito" : 'R54',
        "guantes_dielectricos" : 'R55',
        "linterna" : 'R56',
        "escalera_extension" : 'AB52',
        "escalera_tijera" : 'AB53',
        "escalera_2_pasos" : 'AB54',
        "eslinga_y" : 'AB55',
        "eslinga_posicionamiento" : 'AB56'
    }
    
    wb = load_workbook(filename=Name_file)
    sheet = wb['Hoja1']
    
    for field, value in form.items():
        if field in Cell_assignment:
            cell = Cell_assignment[field]
            sheet[cell] = value
            
    wb.save(Name_file)