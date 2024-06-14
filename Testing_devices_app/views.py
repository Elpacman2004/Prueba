from django.shortcuts import render, redirect
from .forms import GeneralDataDVForm, DFADForm, IMC_form
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image
import os
from PIL import Image as PilImage
import cv2
import numpy as np
from django.http import JsonResponse
from .models import GeneralDataDV



def Index (request):
    return render(request, 'Index TD.html')

def General_data_device (request):
    if request.method == 'POST':
        Form = GeneralDataDVForm(request.POST)
        if Form.is_valid():
            Cleaned = Form.cleaned_data
            print (Cleaned)
            
            New_file_name = f'C:/Users/Dagelec LTDA/Desktop/Pruebas_excel/Mantenimiento/{Cleaned["NameD"]} Hoja de vida.xlsx'
            request.session['file name'] = New_file_name
            
            shutil.copy('C:/Users/Dagelec LTDA/Desktop/Pruebas_excel/Plantilla hojas de vida.xlsx', New_file_name)
            
            wb = load_workbook(filename= New_file_name)
            sheet = wb['Hoja1']
            sheet.title = Cleaned['NameD']
            request.session['Name_sheet'] = sheet.title
            
            sheet['B5'] = Cleaned['NameD']
            sheet['B7'] = Cleaned['Brand']
            sheet['C8'] = Cleaned['Acquisition_Date']
            sheet['C9'] = Cleaned['Supply_Voltage']
            sheet['G7'] = Cleaned['RefModel']
            sheet['G8'] = Cleaned['Supplier']
            sheet['M7'] = Cleaned['Serial_number']
            sheet['M8'] = Cleaned['Invoice']
            if Cleaned['Type_device'] == 'Equipment':
                sheet['H5'] = 'X'
            elif Cleaned['Type_device'] == 'Standard':
                sheet['J5'] = 'X'
            else:
                sheet['N5'] = 'X'
                
            wb.save(New_file_name)
            Form.save()
            
            return redirect('DFAD')
    else:
        Form = GeneralDataDVForm()
        return render(request, 'General_data_DV.html', {'form': Form}) 

def DFAD (request):
    
    if request.method == 'POST':
        Form = DFADForm(request.POST, request.FILES)
        if Form.is_valid():
            File_name = request.session.get('file name')
            Sheet_name = request.session.get('Name_sheet')
            print (Sheet_name)
            
            Cleaned = Form.cleaned_data
            print (Cleaned)
            
            wb = load_workbook(filename= File_name)
            sheet = wb[Sheet_name]  
            sheet['A11'] = Cleaned['Equipment_description']
            img = Cleaned['Photograph']
            with open('temp.png', 'wb+') as temp_file:
                for chunk in img.chunks():
                    temp_file.write(chunk)
                    
            img = cv2.imread('temp.png')
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            _, thresh = cv2.threshold(gray, 240, 255, cv2.THRESH_BINARY)
            contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            mask = np.zeros_like(img)
            cv2.drawContours(mask, contours, -1, (255,255,255), thickness=cv2.FILLED)
            result = cv2.bitwise_and(img, mask)
            cv2.imwrite('temp_no_bg.png', result)
            
            img = PilImage.open('temp_no_bg.png')
            img = img.resize((320, 260))
            img.save('temp_no_bg.png')
            img = Image('temp_no_bg.png')
            sheet.add_image(img, 'H11')
            
            row_number = 26
            for key, value in request.POST.items():
                if key.startswith('dynamicInput'):
                    cell_ref_I = 'A' + str(row_number)
                    cell_ref_U = 'B' + str(row_number)
                    cell_ref_Q = 'C' + str(row_number)
                    cell_ref_D = 'D' + str(row_number)
                    sheet[cell_ref_D] = value

                    counter = key.replace('dynamicInput', '')
                    quantity_key = 'quantity' + counter
                    item_number_key = 'itemNumber' + counter

                    if quantity_key in request.POST:
                        sheet[cell_ref_Q] = request.POST[quantity_key]
                    if item_number_key in request.POST:
                        sheet[cell_ref_I] = request.POST[item_number_key]
                    sheet[cell_ref_U] = 'UNID'

                    row_number += 1
            

            sheet['A52'] = Cleaned['Observations']

            wb.save(File_name)

            os.remove('temp.png')
            
            return redirect('Index TD')
        else:
            print (Form.errors)
    else:
        Form = DFADForm()
        return render(request, 'DFAD.html', {'form': Form})



def Inspection_Maintenance_Calibration (request):
    if request.method == 'POST':
        form = IMC_form(request.POST)
        if form.is_valid():
            Cleaned = form.cleaned_data
            Path = f"C:/Users/Dagelec LTDA/Desktop/Pruebas_excel/Mantenimiento/{Cleaned['Device']} Hoja de vida.xlsx"
            wb = load_workbook(filename= Path )
            sheet = wb['Hoja2']
            
            Infiniti = 0
            Row_number =10
            while Infiniti != 1:
                if sheet[f'A{str(Row_number)}'].value is None:
                    
                    Date_border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='hair'), 
                                        bottom=Side(style='thin'))
                    
                    Left_corner_border = Border(left=Side(style='medium'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='hair'), 
                                        bottom=Side(style='medium'))
                    
                    Center_border = Border(left=Side(style='thin'), 
                                        right=Side(style='thin'), 
                                        top=Side(style='hair'), 
                                        bottom=Side(style='medium'))
                    
                    Right_corner_border = Border(left=Side(style='thin'), 
                                        right=Side(style='medium'), 
                                        top=Side(style='hair'), 
                                        bottom=Side(style='medium'))

                    sheet[f'A{Row_number}'].border = Date_border
                    sheet[f'B{Row_number}'].border = Left_corner_border
                    for col in ['C','D','E', 'F', 'G', 'H', 'I', 'J', 'K',]:
                        sheet[f'{col}{Row_number}'].border = Center_border
                    sheet[f'L{Row_number}'].border = Right_corner_border
                    
                    sheet.merge_cells(f'B{Row_number}:D{Row_number}')
                    sheet.merge_cells(f'H{Row_number}:I{Row_number}')
                    sheet.merge_cells(f'L{Row_number}:M{Row_number}')
                    
                    
                    sheet[f'A{str(Row_number)}'] = Cleaned['Date']
                    sheet[f'B{str(Row_number)}'] = Cleaned['Activity_description']
                    sheet[f'E{str(Row_number)}'] = Cleaned['Activity_type']
                    sheet[f'F{str(Row_number)}'] = Cleaned['Certificate_docment_number']
                    sheet[f'G{str(Row_number)}'] = Cleaned['Satisfactory_result']
                    sheet[f'H{str(Row_number)}'] = Cleaned['Action_to_take']
                    sheet[f'J{str(Row_number)}'] = Cleaned['Next_revew']
                    sheet[f'K{str(Row_number)}'] = Cleaned['Frequency']
                    sheet[f'L{str(Row_number)}'] = Cleaned['Responsible']
                    wb.save(Path)    
                    break
            
                else:
                    Row_number += 1
                    continue    

            return redirect('Index TD')
            
            
        else:
            if 'Device' in form.errors:
                Error = 'El equipo con el mombre {} no existe.'.format(form.data['Device'])
                print(form.errors['Device'])
            return render(request, 'IMC.html', {'form': form,
                                                        'Error': Error
                                                        })
    else: 
        Form = IMC_form ()
        return render(request, 'IMC.html', {'form': Form})

def Search(request):
    q = request.GET.get('q', '')
    Device = GeneralDataDV.objects.filter(NameD__icontains=q)
    results = [Device.NameD for Device in Device]
    return JsonResponse(results, safe=False)