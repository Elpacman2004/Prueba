from django.shortcuts import render, redirect
from .forms import General_data_form, Docments_form, Tecnical_specifications_form
from .models import General_data, Docments as Docs, Tecnical_specifications
from openpyxl import load_workbook
import os
import shutil
from openpyxl.drawing.image import Image

def IndexHVV (request):
    return render(request, 'IndexHVV.html')

def General_Data(request):
    if request.method == 'POST':
        form = General_data_form(request.POST)
        if form.is_valid():
            Cleaned= form.cleaned_data
            print (Cleaned)
            
            folder_path = f"C:/Users/Dagelec LTDA/Documents/Eco_forms/Pruebas_excel/02. HOJAS DE VIDA VEHICULOS/{str(Cleaned['License_plate'])}"
            folder_Service_History = f"C:/Users/Dagelec LTDA/Documents/Eco_forms/Pruebas_excel/02. HOJAS DE VIDA VEHICULOS/{str(Cleaned['License_plate'])}/Hoja de vida"
            folder_preoperational = f"C:/Users/Dagelec LTDA/Documents/Eco_forms/Pruebas_excel/02. HOJAS DE VIDA VEHICULOS/{str(Cleaned['License_plate'])}/Preoperacionales"
            folder_images = f"C:/Users/Dagelec LTDA/Documents/Eco_forms/Pruebas_excel/02. HOJAS DE VIDA VEHICULOS/{str(Cleaned['License_plate'])}/Preoperacionales/Archivos de imagenes"
            os.makedirs(folder_path, exist_ok=True)
            os.makedirs(folder_Service_History, exist_ok=True)
            os.makedirs(folder_preoperational, exist_ok=True)
            os.makedirs(folder_images, exist_ok=True)
            
            File_path = f"C:/Users/Dagelec LTDA/Documents/Eco_forms/Pruebas_excel/02. HOJAS DE VIDA VEHICULOS/{str(Cleaned['License_plate'])}/Hoja de vida/{str(Cleaned['License_plate'])}.xlsx"
            shutil.copy('C:/Users/Dagelec LTDA/Documents/Eco_forms/Pruebas_excel/PlantillaHVV.xlsx', File_path)
            
            request.session['File_path'] = File_path
            
            wb = load_workbook(File_path)
            sheet = wb['Hoja1']
            
            sheet['C5'] = Cleaned['License_plate']
            sheet['C6'] = Cleaned['Class']  
            sheet['C7'] = Cleaned['Engine_number']
            sheet['C8'] = Cleaned['Chassis_number']
            sheet['G5'] = Cleaned['Brand']
            sheet['G6'] = Cleaned['Model']
            sheet['G7'] = Cleaned['Sevice'] 
            sheet['G8'] = Cleaned['Owner']
            sheet['M5'] = Cleaned['Line']
            sheet['M6'] = Cleaned['Body']
            sheet['M7'] = Cleaned['Color']
            sheet['M8'] = Cleaned['Docment']
            
            
            
            wb.save(File_path)
            general_data = form.save()
            request.session['general_data_id'] = general_data.id
            return redirect('Docments')
    else:
        Error = request.session.get('error')
        form = General_data_form()
        return render(request, 'General_data.html', {'form': form,
                                                     'Error': Error })

def Docments(request):
    if request.method == 'POST':
        form = Docments_form(request.POST, request.FILES)
        if form.is_valid():
            Cleaned= form.cleaned_data
            
            general_data_id = request.session['general_data_id']
            
            instance = form.save(commit=False)
            instance.id_License_plate_id = general_data_id
            
            
            File_path = request.session['File_path']
            if File_path == None:
                request.session['error'] = 'Deves llenar el formulario General HVV primero'         
                return redirect ('General HVV')
            wb = load_workbook(File_path)
            sheet = wb['Hoja1']
                
            def get_value_or_na(dictionary, key):
                value = dictionary.get(key)
                if value is None or value == '':
                    return 'N/A'
                return value
            
            sheet['D10'] = get_value_or_na(Cleaned, 'Diving_license')
            sheet['D11'] = get_value_or_na(Cleaned, 'Registration')
            sheet['D12'] = get_value_or_na(Cleaned, 'Gas_sistem_inspection')
            sheet['I10'] = get_value_or_na(Cleaned, 'SOAT')
            sheet['I11'] = get_value_or_na(Cleaned, 'Validity_SOAT')
            sheet['I12'] = get_value_or_na(Cleaned, 'Tecno_mechanical')
            sheet['P10'] = get_value_or_na(Cleaned, 'Operatin_card')
            sheet['P11'] = get_value_or_na(Cleaned, 'Operator')
            sheet['P12'] = get_value_or_na(Cleaned, 'Vehicle_insurance')
            sheet['P13'] = get_value_or_na(Cleaned, 'Policy_number')
            sheet['P14'] = get_value_or_na(Cleaned, 'Vality_Inssurance')
            
            Image_fields = ['Front_photo', 'Side_photo']
            image_cells = ['A16', 'I16']
            
            for field, cell in zip(Image_fields, image_cells):
                image_file = request.FILES[field]
                temp_image_path = f'temp_{field}.png'
                with open(temp_image_path, 'wb+') as destination:
                    for chunk in image_file.chunks():
                        destination.write(chunk)

                img = Image(temp_image_path)

                img.width = 460
                img.height = 322

                sheet.add_image(img, cell)
            
            wb.save(File_path)
            instance.save()
            form.save()
            return redirect('Tecnical')
        
    else:
        form = Docments_form()
        return render(request, 'Docments.html', {'form': form})

def Tecnical_specifications(request):
    if request.method == 'POST':
        form = Tecnical_specifications_form(request.POST)
        if form.is_valid():
            Cleaned= form.cleaned_data
            
            general_data_id = request.session['general_data_id']
            
            instance = form.save(commit=False)
            instance.id_License_plate_id = general_data_id
            
            File_path = request.session['File_path']
            if File_path == None:
                request.session['error'] = 'Deves llenar el formulario General HVV primero'         
                return redirect ('General HVV')
            wb = load_workbook(File_path)
            sheet = wb['Hoja1']
            sheet['F18'] = Cleaned['Fuel']
            sheet['F19'] = Cleaned['Drivetrain']
            sheet['K18'] = Cleaned['Displacement']
            sheet['K19'] = Cleaned['Drive_type']
            sheet['P18'] = Cleaned['Tires']
            sheet['P19'] = f"{Cleaned['Load_capacity']}PSJ"
            
            sheet['D21'] = Cleaned['Air_conditioning']
            sheet['D22'] = Cleaned['Anti_lock_Braking_system']
            sheet['D23'] = Cleaned['Electronic_Brake_Distribution']
            sheet['D24'] = Cleaned['Brake_assist']
            sheet['D25'] = Cleaned['Third_brake_light']
            sheet['D26'] = Cleaned['Side_steps']
            sheet['N21'] = Cleaned['Airbag']    
            sheet['N22'] = Cleaned['Fog_lights']
            sheet['N23'] = Cleaned['Central_locking']
            sheet['N24'] = Cleaned['Seat_covers']
            sheet['N25'] = Cleaned['Radio']
            sheet['N26'] = Cleaned['Power_windows']
            
            sheet['A28'] = f"OBSERVACIONES GENERALES: {Cleaned['General_observatios']}"
            
            wb.save(File_path)
            instance.save()
            form.save()
            return redirect('InexHVV')
        
    form = Tecnical_specifications_form()
    return render(request, 'Tecnical_specifications.html', {'form': form})