from django.shortcuts import render, redirect
from .froms import GeneralDataForm, Staff_requisitionForm, Data_of_selected_personnelForm, Conditions_of_employmentForm, Salary_assignment_form
import os
import shutil
from openpyxl import Workbook, load_workbook
import base64
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.core.files.base import ContentFile
from django.contrib import messages
from PIL import Image
import io
import time

def IndexSPV (request):
    return render(request, 'IndexSPV.html')

def GeneralForm_SPV (request):
    if request.method == 'POST':
        form = GeneralDataForm(request.POST)
        if form.is_valid():
            Clean_data = form.cleaned_data
            print(Clean_data)
            New_file =  f'C:/Users/Dagelec LTDA/Desktop/Pruebas_excel/SPV/{Clean_data["Name"]} SPV.xlsx'
            request.session['file name'] = New_file
            
            shutil.copy('C:/Users/Dagelec LTDA/Desktop/Pruebas_excel/PlantillaSPV.xlsx', New_file)
            
            wb = load_workbook(filename= New_file)
            
            sheet = wb['F-S4-01 Rev 03']
            sheet['B7'] = Clean_data['City']
            sheet['B8'] = Clean_data['Name']
            sheet['B9'] = Clean_data['Process_Project']
            sheet['N7'] = Clean_data['Date_of_application']
            sheet['N8'] = Clean_data['Applicants_Position']
            
            wb.save(New_file)
            form.save()
            return redirect ('Staff requisition')
        
    else:
        error = request.session.get('error', None)
        print (error)
        form = GeneralDataForm()
        return render(request, 'Forms/GeneralForm.html', {
                                                        'form': form,
                                                        'error': error,
                                                          })
    
def Staff_requisition (request):
    if request.method == 'POST':
        form = Staff_requisitionForm(request.POST)
        if form.is_valid():
            file_name = request.session.get('file name', None)
            
            Clean_data = form.cleaned_data
            print(file_name)
            print(Clean_data)
            
            
            wb = load_workbook(filename= file_name)
            sheet = wb['F-S4-01 Rev 03']
            if Clean_data['Reason'] == 'New_Position':
                sheet['C13'] = 'X'
            elif Clean_data['Reason'] == 'Temporary':
                sheet['E13'] = 'X'
            elif Clean_data['Reason'] == 'New_Project':
                sheet['G13'] = 'X'  
            elif Clean_data['Reason'] == 'Other':
                sheet['I13'] = 'X'
            else:
                print ('Error value is not valid')
            
            if Clean_data['Which'] == None:
                sheet['L13'] = ''
            else:
                sheet['L13'] = Clean_data['Which']
            
            sheet['B14'] = Clean_data['Name_of_the_Position_Requested']
            sheet['B15'] = Clean_data['Job_Profile']
                
            wb.save(file_name)
            
            return redirect ('Data personnel')
        
        else:
            print(form.errors)
            return render(request, 'Forms/Staff requisition form.html', {'form': form})
            
    else:
        form = Staff_requisitionForm()
        return render(request, 'Forms/Staff requisition form.html', {'form': form})
    
def Data_of_selected_personnel (request):
        if request.method == 'POST':
            form = Data_of_selected_personnelForm(request.POST)
            if form.is_valid():
                Name = request.session.get('file name')
                if Name is None:
                    request.session['error'] = 'Deves llenar el formulario General SPV primero'         
                    return redirect ('General SPV') 
                
                
                Clean_data = form.cleaned_data
                print(Clean_data)
                
                wb = load_workbook(filename= Name)
                sheet = wb['F-S4-01 Rev 03']
            
                sheet['B20'] = Clean_data['Name'] 
                sheet['K20'] = Clean_data['Identification_document']
                sheet['B21'] = Clean_data['Adress']
                if Clean_data['Phone_number'] == 'None':
                    sheet['B23'] = ''
                else:
                    sheet['B22'] = Clean_data['Phone_number']
                sheet['N21'] = Clean_data['Movile_number']
                sheet['B22'] = Clean_data['Current_HIP']
                sheet['G22'] = Clean_data['AFP']
                sheet['L22'] = Clean_data['RH']
                if Clean_data['Yellow_fever'] == 'Yes':
                    sheet['K23'] = 'X'
                else:
                    sheet['M23'] = 'X'
                if Clean_data['Tetanus'] == 'Yes':
                    sheet['R23'] = 'X'
                else:
                    sheet['T23'] = 'X'
                if Clean_data['Medical_examination'] == 'Yes':
                    sheet['C25'] = 'X'
                else:
                    sheet['E25'] = 'X'
                if Clean_data['Does_the_applicant_meet_the_job_profile'] == 'Yes':
                    sheet['N25'] = 'X'
                elif Clean_data['Does_the_applicant_meet_the_job_profile'] == 'No':
                    sheet['P25'] = 'X'
                else:
                    sheet['T25'] = 'X'
                sheet['B26'] = Clean_data['Pants']
                sheet['G26'] = Clean_data['Shirt']  
                sheet['K26'] = Clean_data['Shoes']
                sheet['R26'] = Clean_data['Overall']
                if Clean_data['Phone_number'] == 'None':
                    pass
                else:
                    sheet['B27'] = Clean_data['Observations']
                if Clean_data['Phone_number'] == 'None':
                    pass
                else:
                    sheet['B29'] = Clean_data['Work_References']
                sheet['B33'] = Clean_data['Personal_References']
                
                wb.save(Name)
                
                return redirect ('Conditions of employment')
        else:
            form = Data_of_selected_personnelForm()
            return render(request, 'Forms/Data of selected personnel.html', {'form': form})
        
def Conditions_of_employment (request):
    if request.method == 'POST':
        form = Conditions_of_employmentForm(request.POST)
        if form.is_valid():
            Name = request.session.get('file name')
            if Name is None:
                print("No file name in session")
                return redirect('IndexSPV')
            Clean_data = form.cleaned_data
            
            wb = load_workbook(filename= Name)
            sheet = wb['F-S4-01 Rev 03']
            if Clean_data['Hired'] == 'Yes':
                sheet['C39'] = 'X'
            else:
                sheet['E39'] = 'X'
            sheet['H39'] = Clean_data['Type_of_Contract']
            if Clean_data['Overtime_Hours'] == 'Yes':
                sheet['Q39'] = 'X'
            else:
                sheet['T39'] = 'X'
            sheet['B40'] = Clean_data['Duration_of_Engagement']
            if Clean_data['Management_and_Trust'] == 'Yes':
                sheet['G40'] = 'X'
            else:
                sheet['I40'] = 'X'
            sheet['O40'] = Clean_data['Workplace']
            sheet['B41'] = Clean_data['Start_Date']
            
            Monday_to_Friday = f'El horario de lunes a viernes es de {Clean_data["Monday_to_Friday_start"]} a {Clean_data["Monday_to_Friday_end"]}'
            Saturday = f'El horario de los sabados es de {Clean_data["Saturday_start"]} a {Clean_data["Saturday_end"]}'
            
            sheet['J41'] = Monday_to_Friday
            sheet['Q41'] = Saturday
            
            wb.save(Name)
            
            
            
            return redirect ('Salary assignment')
        else:
            print(form.errors)
            return render(request, 'Forms/Conditions of employment.html', {'form': form})
    
    else:
        form = Conditions_of_employmentForm()
        return render(request, 'Forms/Conditions of employment.html', {'form': form})

def Salary_assignment (request):
    if request.method == 'POST':
        form = Salary_assignment_form(request.POST)
        if form.is_valid():
            Name = request.session.get('file name')
            if Name is None:
                print("No file name in session")
                return redirect('IndexSPV')
            
            Clean_data = form.cleaned_data
            print(Clean_data)
            wb = load_workbook(filename= Name)
            sheet = wb['F-S4-01 Rev 03']

            Value = f'$ {str(Clean_data["Basic_salary"])}'    
            sheet['B44'] = Value
                
            if Clean_data['Transportation_Allowance'] == 'Yes':
                sheet['I44'] = 'X'
                Value = f'$ {str(Clean_data["Value_Transportation_Allowance"])}' 
                sheet['L44'] = Value
            else:
                sheet['K44'] = 'X'

            if Clean_data['Bonus'] == None:
                sheet['R44'] = '$'
            else:
                Value = f'$ {str(Clean_data["Bonus"])}' 
                sheet['R44'] = Value
            
            if Clean_data['Travel_Allowance'] == 'Yes':
                sheet['C46'] = 'X'
                Value = f'$ {str(Clean_data["Value_Travel_Allowance"])}'
                sheet['F46'] = Value
            else:
                sheet['E46'] = 'X'
            
            if Clean_data['Other'] == 'Yes':
                sheet['J46'] = 'X'
                sheet['O46'] = Clean_data['Which']
            else:
                sheet['L46'] = 'X'  

            sheet['B47'] = Clean_data['Observations']
            
            wb.save(Name)
  
            return redirect ('Signatures')
        else:
            print(form.errors)
            return render(request, 'Forms/Conditions of employment.html', {'form': form})

    else:
        form = Salary_assignment_form()
        return render(request, 'Forms/Salary assignment.html', {'form': form})
    
def Signatures (request):
    if request.method == 'POST':
        Name = request.session.get('file name')
        
        if Name is None:
                print("No file name in session")
                return redirect('IndexSPV')
        
        signatures_data = [
            request.POST.get('signature1'),
            request.POST.get('signature2'),
            request.POST.get('signature3'),
            request.POST.get('signature4'),
            request.POST.get('signature5'),
            request.POST.get('signature6'),
        ]
        
        cells = ['B16', 'H16', 'N16', 'A51', 'F51', 'L51']

        wb = load_workbook(filename= Name)
        sheet = wb['F-S4-01 Rev 03']
        
        open_files = set()

        for i, signature_data in enumerate(signatures_data, start=1):
            if ';base64,' in signature_data:
                format, imgstr = signature_data.split(';base64,')
                ext = format.split('/')[-1]
                data = ContentFile(base64.b64decode(imgstr), name='temp.' + ext)

                # Guarda los datos en un BytesIO para poder leerlos varias veces
                data_io = io.BytesIO(data.read())

                image = Image.open(data_io)

                if image.getbbox():
                    with open(f'signature{i}.png', 'wb') as f:
                        # Retrocede el puntero del BytesIO al principio
                        data_io.seek(0)
                        f.write(data_io.read())

                    # Retrocede el puntero del BytesIO al principio
                    data_io.seek(0)
                    img = Image.open(data_io)

                    # Cambia el tamaño de la imagen
                    img = img.resize((130, 50))

                    # Guarda la imagen en un archivo
                    img.save(f'signature{i}.png')

                    # Abre el archivo de imagen y lo agrega al conjunto de archivos abiertos
                    img_file = open(f'signature{i}.png', 'rb')
                    open_files.add(img_file)

                    # Carga la imagen con openpyxl
                    img = OpenpyxlImage(img_file)

                    sheet.add_image(img, cells[i-1])
                else:
                    messages.error(request, 'Debes firmar todos los campos')
                    return redirect('Signatures')

        wb.save(Name)

        # Cierra todos los archivos abiertos
        for img_file in open_files:
            img_file.close()

        for i in range(1, len(signatures_data) + 1):
            if os.path.exists(f'signature{i}.png'):
                os.remove(f'signature{i}.png')

        return redirect('Index')

    return render(request, 'Forms/Signatures.html')